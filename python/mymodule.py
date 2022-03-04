import csv
import openpyxl as xl

def read_csv(fname):
    """
    A helper function to read csv files. 
    """
    out = []
    with open(fname, 'r') as f: #open the file
        reader = csv.reader(f) #create the reader object
        for row in reader: #for each row in the file
            out.append(row) #add it to the output
    return out

def read_xlsx(fname, num_rows=None):
    """
    A helper function to read XSLX files. 
    """
    out = []
    wb = xl.load_workbook( filename = fname) #create workbook object
    sheet = wb.active #currently only reads the active sheet
    if type(num_rows) == int: #num_rows allows us to select how many rows we need. Useful for testing functions on a subsection of the data rather than waiting for the whole thing to load. 
        last_row = num_rows
    else:
        last_row = sheet.max_row
    for i in range(1, last_row+1): #loop through the rows
        row = []
        for j in range(1, sheet.max_column+1): #loop through the columns
            cell = sheet.cell(row = i, column = j) #add the column to the rows
            row.append(cell.value)
        out.append(row) #add the rows to the output

    return out

import csv
import openpyxl as xl

def read_csv(fname):
    """
    A helper function to read csv files. 
    """
    out = []
    with open(fname, 'r') as f:
        reader = csv.reader(f)
        for row in reader:
            out.append(row)
    return out

def read_xlsx(fname, num_rows=None):
    """
    A helper function to read XSLX files. 
    """
    out = []
    wb = xl.load_workbook( filename = fname)
    sheet = wb.active #currently only reads the active sheet
    if type(num_rows) == int:
        last_row = num_rows
    else:
        last_row = sheet.max_row
    for i in range(1, last_row+1):
        row = []
        for j in range(1, sheet.max_column+1):
            cell = sheet.cell(row = i, column = j)
            row.append(cell.value)
        out.append(row)

    return out
